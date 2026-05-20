#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
批量更新脚本
读取用户填写的Excel（含上传后的FileID），匹配书名后批量更新云数据库

使用方法:
  python batch_update.py --input my_upload.xlsx

Excel格式要求（必须有表头）:
  书名 | 音频FileID | 封面FileID
"""

import os
import sys
import json
import httpx
import logging
from pathlib import Path
from typing import List, Dict
from datetime import datetime

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
logger = logging.getLogger(__name__)


# ==================== 配置 ====================
SCRIPT_DIR = Path(__file__).parent


def load_env():
    """加载环境变量"""
    env_path = SCRIPT_DIR / ".env"
    if env_path.exists():
        with open(env_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    os.environ.setdefault(key.strip(), value.strip())


def get_access_token() -> str:
    """获取微信access_token"""
    app_id = os.getenv("WECHAT_APP_ID")
    secret = os.getenv("WECHAT_SECRET")

    url = f"https://api.weixin.qq.com/cgi-bin/token?grant_type=client_credential&appid={app_id}&secret={secret}"
    with httpx.Client(timeout=30.0) as client:
        response = client.get(url)
        result = response.json()
        if "access_token" in result:
            return result["access_token"]
        raise RuntimeError(f"获取access_token失败: {result}")


def query_book_by_title(title: str) -> Dict:
    """根据书名查询书籍记录"""
    access_token = get_access_token()
    env_id = os.getenv("WECHAT_ENV_ID")

    query = f"""
db.collection('books').where({{title: '{title}'}}).limit(1).get()
"""
    url = f"https://api.weixin.qq.com/tcb/databasequery?access_token={access_token}"
    data = {"env": env_id, "query": query.strip()}

    with httpx.Client(timeout=30.0) as client:
        response = client.post(url, json=data)
        result = response.json()

        if result.get("errcode") == 0:
            data_list = result.get("data", [])
            if data_list:
                return json.loads(data_list[0])
    return {}


def update_book_urls(book_id: str, audio_url: str = None, cover_url: str = None) -> bool:
    """更新书籍的audioUrl和coverUrl"""
    access_token = get_access_token()
    env_id = os.getenv("WECHAT_ENV_ID")

    # 构建更新字段
    updates = []
    if audio_url:
        updates.append(f'audioUrl: "{audio_url}"')
    if cover_url:
        updates.append(f'coverUrl: "{cover_url}"')

    if not updates:
        logger.warning("没有需要更新的字段")
        return False

    query = f"""
db.collection('books').doc('{book_id}').update({{data: {{{', '.join(updates)}}}}})
"""
    url = f"https://api.weixin.qq.com/tcb/databaseupdate?access_token={access_token}"
    data = {"env": env_id, "query": query.strip()}

    with httpx.Client(timeout=30.0) as client:
        response = client.post(url, json=data)
        result = response.json()

        if result.get("errcode") == 0:
            logger.info(f"✅ 更新成功: {book_id}")
            return True
        else:
            logger.error(f"更新失败: {result}")
            return False


def load_user_file(file_path: Path) -> List[Dict]:
    """读取用户填写的Excel/CSV"""
    ext = file_path.suffix.lower()

    if ext == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # 读取表头
        headers = [cell.value for cell in ws[1]]
        logger.info(f"表头: {headers}")

        # 查找列索引
        col_title = None
        col_audio = None
        col_cover = None

        for i, h in enumerate(headers):
            h_lower = str(h).lower()
            if "书名" in h_lower or "title" in h_lower:
                col_title = i
            elif "音频" in h_lower or "audio" in h_lower:
                col_audio = i
            elif "封面" in h_lower or "cover" in h_lower:
                col_cover = i

        if col_title is None:
            raise ValueError("Excel缺少'书名'列")

        # 读取数据行
        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[col_title]:  # 有书名
                records.append({
                    "书名": row[col_title],
                    "音频FileID": row[col_audio] if col_audio else "",
                    "封面FileID": row[col_cover] if col_cover else ""
                })
        return records

    elif ext == ".csv":
        import csv
        records = []
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                records.append(row)
        return records

    else:
        raise ValueError(f"不支持的文件格式: {ext}")


def main():
    import argparse
    parser = argparse.ArgumentParser(description="批量更新云数据库URL")
    parser.add_argument("--input", type=str, required=True, help="用户填写的Excel/CSV文件")
    parser.add_argument("--dry-run", action="store_true", help="仅模拟，不实际更新")
    args = parser.parse_args()

    # 加载配置
    load_env()

    # 读取用户文件
    input_path = Path(args.input)
    if not input_path.exists():
        logger.error(f"文件不存在: {input_path}")
        return

    records = load_user_file(input_path)
    logger.info(f"共读取 {len(records)} 条记录")

    # 遍历更新
    success_count = 0
    fail_count = 0
    skip_count = 0

    for record in records:
        title = record.get("书名", "").strip()
        audio_url = record.get("音频FileID", "").strip()
        cover_url = record.get("封面FileID", "").strip()

        if not title:
            skip_count += 1
            continue

        # 查询书籍
        book = query_book_by_title(title)
        if not book:
            logger.warning(f"❌ 云端未找到: {title}")
            fail_count += 1
            continue

        book_id = book.get("_id", "")
        if args.dry_run:
            logger.info(f"[模拟] {title}: audio={audio_url[:30] if audio_url else 'N/A'}..., cover={cover_url[:30] if cover_url else 'N/A'}...")
            success_count += 1
            continue

        # 更新
        if update_book_urls(book_id, audio_url or None, cover_url or None):
            logger.info(f"✅ 更新成功: {title}")
            success_count += 1
        else:
            logger.error(f"❌ 更新失败: {title}")
            fail_count += 1

    logger.info(f"\n完成: 成功={success_count}, 失败={fail_count}, 跳过={skip_count}")


if __name__ == "__main__":
    main()