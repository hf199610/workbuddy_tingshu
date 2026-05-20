#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
云数据库导出脚本
导出云数据库books集合的所有记录到Excel

使用方法:
  python export_db.py                    # 导出全部
  python export_db.py --limit 10          # 只导出前10条
  python export_db.py --output books.xlsx  # 指定输出文件
"""

import os
import sys
import json
import httpx
import logging
from pathlib import Path
from typing import List, Dict, Optional
from datetime import datetime

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
logger = logging.getLogger(__name__)


# ==================== 配置 ====================
SCRIPT_DIR = Path(__file__).parent
OUTPUT_DIR = SCRIPT_DIR / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# 输出文件默认用当天日期
DEFAULT_OUTPUT = OUTPUT_DIR / f"db_export_{datetime.now().strftime('%Y%m%d')}.xlsx"

# 数据库字段列表
DB_FIELDS = [
    "_id", "title", "author", "category", "categoryName",
    "publisher", "isbn", "publishDate", "pages",
    "description", "coverUrl", "coverColor",
    "script", "cleanScript", "quotes",
    "audioFileId", "audioUrl", "updateTime"
]


def load_env():
    """加载环境变量"""
    env_path = SCRIPT_DIR / ".env"
    if env_path.exists():
        with open(env_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    os.environ.setdefault(key.strip(), value.strip())


def get_access_token() -> str:
    """获取微信access_token"""
    app_id = os.getenv("WECHAT_APP_ID")
    secret = os.getenv("WECHAT_SECRET")

    url = f"https://api.weixin.qq.com/cgi-bin/token?grant_type=client_credential&appid={app_id}&secret={secret}"
    with httpx.Client(timeout=30.0) as client:
        response = client.get(url)
        result = response.json()
        if "access_token" in result:
            logger.info("✅ 获取access_token成功")
            return result["access_token"]
        raise RuntimeError(f"获取access_token失败: {result}")


def query_all_books(limit: int = 1000) -> List[Dict]:
    """查询云数据库所有书籍记录"""
    access_token = get_access_token()
    env_id = os.getenv("WECHAT_ENV_ID")

    # 查询所有记录（wx-server API限制最多100条，需要分页）
    all_books = []
    skip = 0
    page_size = 100

    while skip < limit:
        query = f"""
db.collection('books').orderBy('updateTime', 'desc').skip({skip}).limit({page_size}).get()
"""
        url = f"https://api.weixin.qq.com/tcb/databasequery?access_token={access_token}"
        data = {"env": env_id, "query": query.strip()}

        with httpx.Client(timeout=30.0) as client:
            response = client.post(url, json=data)
            result = response.json()

            if result.get("errcode") != 0:
                logger.error(f"查询失败: {result}")
                break

            data_list = result.get("data", [])
            if not data_list:
                break

            # 解析每条记录
            for item in data_list:
                try:
                    book = json.loads(item)
                    all_books.append(book)
                except json.JSONDecodeError as e:
                    logger.warning(f"解析失败: {e}")

            logger.info(f"已获取 {len(all_books)} 条记录...")
            skip += page_size

            if len(data_list) < page_size:
                break

    logger.info(f"✅ 共获取 {len(all_books)} 条书籍记录")
    return all_books


def export_to_excel(books: List[Dict], output_path: Path):
    """导出书籍记录到Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "书籍列表"

        # 表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # 写入表头
        headers = ["序号", "书名", "作者", "分类ID", "分类名", "出版社", "ISBN",
                "出版日期", "页数", "简介", "封面URL", "封面颜色",
                "金句(数组)", "音频FileID", "音频URL", "更新时间"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 写入数据
        for row, book in enumerate(books, 2):
            ws.cell(row=row, column=1, value=row-1)  # 序号
            ws.cell(row=row, column=2, value=book.get("title", ""))
            ws.cell(row=row, column=3, value=book.get("author", ""))
            ws.cell(row=row, column=4, value=book.get("category", ""))
            ws.cell(row=row, column=5, value=book.get("categoryName", ""))
            ws.cell(row=row, column=6, value=book.get("publisher", ""))
            ws.cell(row=row, column=7, value=book.get("isbn", ""))
            ws.cell(row=row, column=8, value=book.get("publishDate", ""))
            ws.cell(row=row, column=9, value=book.get("pages", ""))
            ws.cell(row=row, column=10, value=book.get("description", ""))
            ws.cell(row=row, column=11, value=book.get("coverUrl", ""))
            ws.cell(row=row, column=12, value=book.get("coverColor", ""))
            # 金句数组转字符串
            quotes = book.get("quotes", [])
            quotes_str = "|".join(quotes) if quotes else ""
            ws.cell(row=row, column=13, value=quotes_str)
            ws.cell(row=row, column=14, value=book.get("audioFileId", ""))
            ws.cell(row=row, column=15, value=book.get("audioUrl", ""))
            # 时间戳转日期
            update_time = book.get("updateTime", 0)
            if update_time:
                from datetime import datetime
                update_date = datetime.fromtimestamp(update_time/1000)
                ws.cell(row=row, column=16, value=update_date.strftime("%Y-%m-%d %H:%M"))
            else:
                ws.cell(row=row, column=16, value="")

        # 调整列宽
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 8
        ws.column_dimensions['J'].width = 40
        ws.column_dimensions['K'].width = 40
        ws.column_dimensions['L'].width = 10
        ws.column_dimensions['M'].width = 30
        ws.column_dimensions['N'].width = 40
        ws.column_dimensions['O'].width = 40
        ws.column_dimensions['P'].width = 15

        wb.save(output_path)
        logger.info(f"✅ 已导出到: {output_path}")

    except ImportError:
        # 没有openpyxl时导出CSV
        import csv
        with open(output_path.with_suffix('.csv'), 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(["序号", "书名", "作者", "分类", "封面URL", "音频URL"])

            for i, book in enumerate(books, 1):
                writer.writerow([
                    i,
                    book.get("title", ""),
                    book.get("author", ""),
                    book.get("categoryName", ""),
                    book.get("coverUrl", ""),
                    book.get("audioUrl", "")
                ])
        logger.info(f"✅ 已导出到CSV: {output_path.with_suffix('.csv')}")


def main():
    import argparse
    parser = argparse.ArgumentParser(description="导出云数据库书籍记录")
    parser.add_argument("--limit", type=int, default=1000, help="导出数量限制")
    parser.add_argument("--output", type=str, default="", help="输出文件路径")
    args = parser.parse_args()

    # 加载配置
    load_env()

    # 确定输出路径
    if args.output:
        output_path = Path(args.output)
    else:
        output_path = DEFAULT_OUTPUT

    # 导出
    books = query_all_books(args.limit)
    if books:
        export_to_excel(books, output_path)
    else:
        logger.warning("没有找到任何书籍记录")


if __name__ == "__main__":
    main()