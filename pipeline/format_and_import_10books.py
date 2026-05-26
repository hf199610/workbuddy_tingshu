#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""格式化讲解稿并导入10本书到云数据库"""

import os, json, time, logging, httpx, re, csv
from pathlib import Path
from mutagen.mp3 import MP3

PIPELINE_DIR = Path(__file__).parent
OUTPUT_DIR = PIPELINE_DIR / "output"
CSV_FILE = PIPELINE_DIR.parent / "data_source" / "book_list_500_enhanced.csv"
EXCEL_FILE = PIPELINE_DIR.parent / "data_source" / "book_list_500_enhanced.xlsx"

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)


def load_env():
    for env_path in [PIPELINE_DIR / ".env", PIPELINE_DIR.parent / ".env"]:
        if env_path.exists():
            with open(env_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#") and "=" in line:
                        key, value = line.split("=", 1)
                        os.environ.setdefault(key.strip(), value.strip())


def get_access_token():
    app_id = os.environ.get("WECHAT_APP_ID")
    secret = os.environ.get("WECHAT_SECRET")
    resp = httpx.get(f"https://api.weixin.qq.com/cgi-bin/token?grant_type=client_credential&appid={app_id}&secret={secret}", timeout=30)
    return resp.json()["access_token"]


def escape_str(s):
    """JSON字符串转义"""
    s = str(s)
    s = s.replace("\\", "\\\\")
    s = s.replace('"', '\\"')
    s = s.replace("\n", "\\n")
    s = s.replace("\r", "\\r")
    s = s.replace("\t", "\\t")
    s = s.replace('"', '"').replace('"', '"')
    s = s.replace(''', "'").replace(''', "'")
    return s


def format_script(text):
    """格式化讲解稿：清理**、#等朗读不适应的符号"""
    if not text:
        return text

    # 替换 **xxx** 粗体格式 -> xxx
    text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)
    # 替换 *xxx* 斜体格式 -> xxx
    text = re.sub(r'\*(.+?)\*', r'\1', text)
    # 替换 ### xxx 标题格式 -> xxx
    text = re.sub(r'^#{1,6}\s+', '', text, flags=re.MULTILINE)
    # 替换 - xxx 列表格式 -> xxx
    text = re.sub(r'^[\-\*]\s+', '', text, flags=re.MULTILINE)
    # 替换数字. xxx 列表格式
    text = re.sub(r'^\d+\.\s+', '', text, flags=re.MULTILINE)
    # 清理多余的空行
    text = re.sub(r'\n{3,}', '\n\n', text)
    # 清理行首行尾空格
    lines = [line.strip() for line in text.split('\n')]
    text = '\n'.join(lines)
    return text


def generate_quotes(title, author, count=10):
    """生成多条金句"""
    from anthropic import Anthropic
    client = Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"),
                       base_url=os.environ.get("ANTHROPIC_BASE_URL"))

    prompt = f"""请为《{title}》（作者：{author}）生成{count}条经典金句。

请严格按以下JSON数组格式输出（不要包含任何其他内容）：
[
  {{"content": "金句内容1", "author": "作者或人物1"}},
  {{"content": "金句内容2", "author": "作者或人物2"}},
  {{"content": "金句内容3", "author": "作者或人物3"}}
]

要求：
1. 必须生成恰好{count}条金句
2. 每条金句不超过50字
3. 每条必须包含content和author两个字段
4. 只输出JSON数组，不要任何其他说明文字"""

    for attempt in range(3):
        try:
            msg = client.messages.create(
                model=os.environ.get("MINIMAX_MODEL", "MiniMax-M2.7"),
                max_tokens=4000,
                messages=[{"role": "user", "content": prompt}]
            )
            text = ""
            for block in msg.content:
                if hasattr(block, 'type') and block.type == 'text':
                    text = block.text.strip()
                    break
            else:
                text = str(msg.content)

            # 解析JSON
            matches = re.findall(r'\{[^{}]*\}', text)
            if len(matches) >= count:
                quotes = []
                for m in matches[:count]:
                    try:
                        q = json.loads(m)
                        if 'content' in q and 'author' in q:
                            quotes.append(q)
                    except:
                        pass
                if len(quotes) >= count:
                    return quotes

        except Exception as e:
            logger.error(f"第{attempt+1}次异常: {e}")
            time.sleep(5)

    return [{"content": f"《{title}》中的经典句子。", "author": author}]


def gen_audio_male(text, output_path):
    """生成男声音频（分段处理长文本）"""
    import edge_tts

    MAX_CHUNK = 2980
    if len(text) <= MAX_CHUNK:
        return _gen_audio_single(text, output_path)

    # 分段生成
    paragraphs = text.split('\n')
    chunks = []
    current = []
    current_len = 0

    for para in paragraphs:
        para_len = len(para)
        if current_len + para_len > MAX_CHUNK and current:
            chunks.append('\n'.join(current))
            current = []
            current_len = 0
        current.append(para)
        current_len += para_len

    if current:
        chunks.append('\n'.join(current))

    logger.info(f"文本过长，分{len(chunks)}段生成音频")

    # 逐段生成并合并
    from pydub import AudioSegment

    output_path.parent.mkdir(parents=True, exist_ok=True)

    combined = AudioSegment.empty()
    for i, chunk in enumerate(chunks):
        chunk_file = output_path.parent / f"temp_chunk_{i}.mp3"
        result = _gen_audio_single(chunk, chunk_file)
        if result:
            combined += AudioSegment.from_mp3(str(chunk_file))
            chunk_file.unlink(missing_ok=True)
        time.sleep(1)

    combined.export(str(output_path), format="mp3")
    logger.info(f"音频合并完成: {output_path}")
    return output_path


def _gen_audio_single(text, output_path):
    """单个音频生成"""
    import edge_tts
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        communicate = edge_tts.Communicate(text, "zh-CN-YunxiNeural")
        communicate.save_sync(str(output_path))
        logger.info(f"音频生成成功: {output_path}")
        return output_path
    except Exception as e:
        logger.error(f"生成音频失败: {e}")
        return None


def get_audio_duration(path):
    try:
        audio = MP3(str(path))
        total_sec = int(audio.info.length)
        mins, secs = divmod(total_sec, 60)
        hours, mins = divmod(mins, 60)
        return f"{hours:02d}:{mins:02d}:{secs:02d}"
    except:
        return "00:05:00"


def upload_file(access_token, env_id, file_path, cloud_path):
    if not os.path.exists(file_path):
        return None
    step1_url = f"https://api.weixin.qq.com/tcb/uploadfile?access_token={access_token}"
    resp1 = httpx.post(step1_url, json={"env": env_id, "path": cloud_path}, timeout=30)
    result1 = resp1.json()
    if result1.get("errcode") != 0:
        logger.error(f"获取上传链接失败: {result1}")
        return None

    with open(file_path, "rb") as f:
        file_content = f.read()
    files = {
        "key": (None, cloud_path),
        "Signature": (None, result1.get("authorization")),
        "x-cos-security-token": (None, result1.get("token")),
        "x-cos-meta-fileid": (None, result1.get("cos_file_id")),
        "file": (os.path.basename(file_path), file_content),
    }
    resp2 = httpx.post(result1.get("url"), files=files, timeout=180)
    if resp2.status_code in [200, 204]:
        return result1.get("file_id", "")
    logger.error(f"上传失败: {resp2.status_code}")
    return None


def add_book_to_db(access_token, env_id, data):
    data_str = json.dumps(data, ensure_ascii=False)
    add_query = f'db.collection("books").add({{data: {data_str}}})'
    resp = httpx.post(f"https://api.weixin.qq.com/tcb/databaseadd?access_token={access_token}",
                      json={"env": env_id, "query": add_query}, timeout=60)
    result = resp.json()
    if result.get("errcode") == 0:
        return result.get("id_list", [None])[0]
    logger.error(f"导入失败: {result}")
    return None


def update_csv(csv_file, title, db_id, import_time, quotes_count, duration):
    """更新CSV状态"""
    rows = []
    for encoding in ['utf-8', 'gbk', 'gb2312', 'latin1']:
        try:
            with open(csv_file, 'r', encoding=encoding) as f:
                reader = csv.reader(f)
                header = next(reader)
                # 确保表头包含新字段
                if '金句数' not in header and '时长' not in header:
                    header.extend(['金句数', '时长'])
                rows.append(header)
                for row in reader:
                    if len(row) > 0 and row[0] == title:
                        row[2] = '是'  # 是否导入
                        row[8] = db_id  # 数据库ID
                        row[7] = import_time  # 导入时间
                        # 确保行有足够长度
                        while len(row) < 11:
                            row.append('')
                        row[9] = str(quotes_count)  # 金句数
                        row[10] = duration  # 时长
                    # 确保每行有足够的列
                    while len(row) < 11:
                        row.append('')
                    rows.append(row)
            break
        except UnicodeDecodeError:
            continue

    with open(csv_file, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def update_excel(excel_file, title, db_id, import_time, quotes_count, duration):
    """更新Excel状态"""
    try:
        import openpyxl
        if not os.path.exists(excel_file):
            logger.warning(f"Excel文件不存在: {excel_file}")
            return

        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        # 检查表头
        headers = [cell.value for cell in ws[1]]
        if '金句数' not in headers:
            ws.cell(1, len(headers) + 1, '金句数')
        if '时长' not in headers:
            ws.cell(1, len(headers) + 1, '时长')

        # 查找并更新行
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == title:
                row[2].value = '是'  # 是否导入
                row[8].value = db_id  # 数据库ID
                row[7].value = import_time  # 导入时间
                row[9].value = quotes_count  # 金句数
                row[10].value = duration  # 时长
                break

        wb.save(excel_file)
        logger.info(f"Excel已更新: {title}")
    except Exception as e:
        logger.error(f"更新Excel失败: {e}")


# 10本书配置
BOOKS = [
    {
        "title": "红楼梦",
        "author": "曹雪芹",
        "category": "经典名著",
        "coverColor": "#8B0000",
        "intro": "《红楼梦》是清代作家曹雪芹创作的长篇小说，被誉为中国古典小说的巅峰之作。小说以贾、史、王、薛四大家族的兴衰为背景，以贾宝玉、林黛玉、薛宝钗的爱情婚姻故事为主线，描绘了一个庞大而精细的贵族家庭的生活画面。全书规模宏大，结构严谨，语言优美，人物形象生动，是中国文学史上最具影响力的作品之一。",
    },
    {
        "title": "三国演义",
        "author": "罗贯中",
        "category": "经典名著",
        "coverColor": "#B8860B",
        "intro": "《三国演义》是元末明初小说家罗贯中创作的长篇章回体历史演义小说。它描写了从东汉末年到西晋初年之间近105年的历史风云，以描写战争为主，讲述了魏、蜀、吴三国之间的政治和军事斗争。作者将兵法三十六计融于字里行间，既有实据，也有虚构，结构宏伟，情节引人入胜。",
    },
    {
        "title": "水浒传",
        "author": "施耐庵",
        "category": "经典名著",
        "coverColor": "#228B22",
        "intro": "《水浒传》是元末明初施耐庵创作的长篇小说，是中国四大名著之一。小说以宋江领导的梁山泊农民起义为题材，塑造了宋江、李逵、武松、林冲等一百单八将的英雄形象。书中揭示了当时社会矛盾的尖锐，歌颂了农民的斗争精神，具有深刻的思想内容和高超的艺术成就。",
    },
    {
        "title": "西游记",
        "author": "吴承恩",
        "category": "经典名著",
        "coverColor": "#FF8C00",
        "intro": "《西游记》是明代吴承恩创作的长篇神魔小说，是中国四大名著之一。小说讲述了唐僧师徒四人西天取经的故事，塑造了孙悟空、猪八戒、沙僧等经典形象。全书以浪漫主义手法描绘了神奇瑰丽的神话世界，寓意深刻，融佛、道、儒三家思想于一体，流传千古，深受喜爱。",
    },
    {
        "title": "活着",
        "author": "余华",
        "category": "经典名著",
        "coverColor": "#4A4A4A",
        "intro": "《活着》是当代作家余华的代表作，讲述了一个农民福贵悲惨的一生。通过福贵一家的遭遇，深刻揭示了人生的苦难与生存的意义。小说语言简洁有力，以平实的笔触描绘了人生的无奈与坚韧，被誉为20世纪最有影响力的作品之一。",
    },
    {
        "title": "平凡的世界",
        "author": "路遥",
        "category": "经典名著",
        "coverColor": "#8B4513",
        "intro": "《平凡的世界》是作家路遥创作的长篇小说，以中国70年代中期到80年代中期为背景，通过复杂的矛盾纠葛，刻画了当时社会各阶层众多普通人的形象。劳动与爱情、挫折与追求、痛苦与欢乐，日常生活与巨大社会冲突纷繁地交织在一起，深刻展示了普通人在大时代历史进程中所走过的艰难道路。",
    },
    {
        "title": "白鹿原",
        "author": "陈忠实",
        "category": "经典名著",
        "coverColor": "#800000",
        "intro": "《白鹿原》是作家陈忠实的代表作，以陕西关中地区白鹿原上白鹿村为缩影，讲述白姓和鹿姓两大家族祖孙三代的恩怨纷争。从清末民初到二十世纪七八十年代，半个多世纪的生死较量，是一部渭河平原50年变迁的雄奇史诗，一轴中国农村斑斓多彩、触目惊心的长幅画卷。",
    },
    {
        "title": "四世同堂",
        "author": "老舍",
        "category": "经典名著",
        "coverColor": "#6B8E23",
        "intro": "《四世同堂》是老舍创作的长篇小说，描写了抗日战争时期北平小羊圈胡同里祁家四代人的生活。小说以祁家为主，兼顾钱家、冠家等其他人家，真实地记录了日本侵略者的残暴行为，以及北平人民在亡国之际的苦难与抗争，是一部民族灵魂的发现和拷问之作。",
    },
    {
        "title": "骆驼祥子",
        "author": "老舍",
        "category": "经典名著",
        "coverColor": "#CD853F",
        "intro": "《骆驼祥子》是老舍的代表作，以20世纪20年代的北京为背景，讲述了一个洋车夫祥子的故事。祥子老实、健壮、坚忍，最大的梦想不过是拥有一辆自己的车。但在那个黑暗世道里，他的希望一次又一次破灭，最终沦为行尸走肉。小说语言朴实生动，深刻的揭露了旧中国的黑暗，表达了对劳动人民的深切同情。",
    },
    {
        "title": "红高粱",
        "author": "莫言",
        "category": "经典名著",
        "coverColor": "#B22222",
        "intro": "《红高粱》是莫言的代表作，以抗日战争时期的高密东北乡为背景，描写了一群土生土长的农民在面对外敌入侵时的英勇抗争。余占鳌、土匪花脖子都是鲜活的英雄形象。小说开创了寻根文学的新局面，以自由不羁的语言、汪洋恣肆的想象、热烈绵绵的感情，谱写了一段离经叛道的抗日传奇。",
    },
]


def main():
    logger.info("=" * 60)
    logger.info("格式化讲解稿并导入10本书到云数据库")
    logger.info("=" * 60)

    load_env()
    env_id = os.environ.get("WECHAT_ENV_ID")

    for i, book in enumerate(BOOKS, 1):
        logger.info(f"\n{'='*60}")
        logger.info(f"【第{i}/10本】处理: {book['title']}")
        logger.info(f"{'='*60}")

        try:
            title = book["title"]

            # Step 1: 读取并格式化讲解稿
            logger.info(f"\n--- Step 1: 格式化讲解稿 ---")
            script_file = OUTPUT_DIR / "data" / f"{title}_script.txt"
            formatted_file = OUTPUT_DIR / "data" / f"{title}_formatted.txt"

            if script_file.exists():
                with open(script_file, "r", encoding="utf-8") as f:
                    raw_script = f.read()
                formatted_script = format_script(raw_script)

                # 保存格式化后的讲解稿
                with open(formatted_file, "w", encoding="utf-8") as f:
                    f.write(formatted_script)

                # 检查清理了多少特殊符号
                raw_stars = raw_script.count('**')
                logger.info(f"讲解稿已格式化，清除 {raw_stars} 个 ** 符号，{len(raw_script) - len(formatted_script)} 字符")
            else:
                logger.warning(f"未找到讲解稿文件: {script_file}")
                formatted_script = ""
                # 尝试生成新的讲解稿
                from anthropic import Anthropic
                client = Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"),
                                   base_url=os.environ.get("ANTHROPIC_BASE_URL"))
                prompt = f"""请为《{title}》（作者：{book['author']}）生成一段6000-8000字的书籍讲解稿。

要求：
1. 详细生动地介绍这本书的写作背景、情节、人物、主题思想
2. 语言通俗易懂，适合听书场景
3. 字数控制在6000-8000字之间
4. 不要使用任何 ** 或 # 等特殊格式符号
5. 只输出讲解稿内容，不要其他说明"""
                msg = client.messages.create(model=os.environ.get("MINIMAX_MODEL", "MiniMax-M2.7"),
                                              max_tokens=12000, messages=[{"role": "user", "content": prompt}])
                for block in msg.content:
                    if hasattr(block, 'type') and block.type == 'text':
                        formatted_script = block.text.strip()
                        break
                else:
                    formatted_script = str(msg.content)

                with open(formatted_file, "w", encoding="utf-8") as f:
                    f.write(formatted_script)

            script_len = len(formatted_script)
            logger.info(f"格式化后讲解稿: {script_len} 字")

            # Step 2: 生成金句（10条）
            logger.info(f"\n--- Step 2: 生成金句(10条) ---")
            quotes = generate_quotes(title, book["author"], count=10)
            logger.info(f"金句: {len(quotes)} 条")

            # Step 3: 生成男声音频
            logger.info(f"\n--- Step 3: 生成男声音频 ---")
            audio_path = OUTPUT_DIR / "audios" / f"{title}_male.mp3"
            if formatted_script:
                gen_audio_male(formatted_script, audio_path)

            # Step 4: 生成封面
            logger.info(f"\n--- Step 4: 生成封面 ---")
            cover_path = OUTPUT_DIR / "covers" / f"{title}.jpg"
            from PIL import Image, ImageDraw, ImageFont
            try:
                img = Image.new("RGB", (400, 600), color=book["coverColor"])
                draw = ImageDraw.Draw(img)
                draw.rectangle([20, 20, 380, 580], outline="white", width=3)
                try:
                    font = ImageFont.truetype("C:/Windows/Fonts/simhei.ttf", 40)
                except:
                    font = ImageFont.load_default()
                lines = [title[j:j+6] for j in range(0, len(title), 6)]
                y = 250
                for line in lines:
                    bbox = draw.textbbox((0, 0), line, font=font)
                    x = (400 - (bbox[2]-bbox[0])) // 2
                    draw.text((x, y), line, fill="white", font=font)
                    y += 60
                bbox = draw.textbbox((0, 0), book["author"], font=font)
                draw.text(((400-(bbox[2]-bbox[0]))//2, y+20), book["author"], fill="#FFFFFFCC", font=font)
                cover_path.parent.mkdir(parents=True, exist_ok=True)
                img.save(cover_path, "JPEG", quality=95)
                logger.info(f"封面生成成功: {cover_path}")
            except Exception as e:
                logger.error(f"生成封面失败: {e}")

            # Step 5: 上传云存储
            logger.info(f"\n--- Step 5: 上传云存储 ---")
            access_token = get_access_token()
            audio_file_id = upload_file(access_token, env_id, str(audio_path), f"audios/{title}_male.mp3") if audio_path.exists() else None
            cover_file_id = upload_file(access_token, env_id, str(cover_path), f"covers/{title}.jpg") if cover_path and cover_path.exists() else None
            logger.info(f"音频FileID: {audio_file_id}")
            logger.info(f"封面FileID: {cover_file_id}")

            # Step 6: 更新数据库
            logger.info(f"\n--- Step 6: 更新数据库 ---")
            duration = get_audio_duration(audio_path) if audio_path.exists() else "00:05:00"
            timestamp = int(time.time() * 1000)
            import_time = time.strftime("%Y-%m-%d %H:%M:%S")

            data_obj = {
                "title": title,
                "author": book["author"],
                "category": book["category"],
                "intro": escape_str(book["intro"]),
                "script": escape_str(formatted_script) if formatted_script else "",
                "scriptLength": script_len,
                "quotes": quotes,
                "quotesCount": len(quotes),
                "audioFileId": audio_file_id or "",
                "audioUrl": audio_file_id or "",
                "coverUrl": cover_file_id or "",
                "coverColor": book["coverColor"],
                "duration": duration,
                "isGenerated": True,
                "isPublished": True,
                "isHot": False,
                "createTime": timestamp,
                "updateTime": timestamp,
            }

            db_id = add_book_to_db(access_token, env_id, data_obj)
            logger.info(f"数据库ID: {db_id}")

            # Step 7: 更新CSV和Excel
            logger.info(f"\n--- Step 7: 更新CSV和Excel ---")
            if db_id:
                update_csv(CSV_FILE, title, db_id, import_time, len(quotes), duration)
                update_excel(EXCEL_FILE, title, db_id, import_time, len(quotes), duration)
                logger.info(f"文档已更新")

            logger.info(f"\n✅ 完成: {title}")
            logger.info(f"   讲解稿: {script_len} 字 (已格式化)")
            logger.info(f"   金句: {len(quotes)} 条")
            logger.info(f"   音频时长: {duration}")

            if i < len(BOOKS):
                logger.info("\n等待30秒后处理下一本...")
                time.sleep(30)

        except Exception as e:
            logger.error(f"处理 {book.get('title', '未知')} 时出错: {e}")
            import traceback
            traceback.print_exc()
            continue

    logger.info("\n" + "=" * 60)
    logger.info("全部10本书处理完成!")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
