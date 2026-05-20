#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
听书小程序 - 全自动Pipeline
整合：书籍信息获取 → MiniMax内容生成 → edge-tts音频 → 字幕清理 → 云端上传

使用方法:
  python full_pipeline.py                          # 全流程（选1本未导入的新书）
  python full_pipeline.py --count 5               # 处理5本
  python full_pipeline.py --book "活着"           # 指定书籍
  python full_pipeline.py --book "活着" --force    # 强制重新处理
"""

import os
import re
import sys
import json
import time
import asyncio
import logging
import argparse
import httpx
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional
from urllib.parse import quote as url_quote

# 路径配置
BASE_DIR = Path(__file__).parent.parent
DATA_DIR = BASE_DIR / "data_source"
OUTPUT_DIR = Path(__file__).parent / "output"
AUDIO_DIR = OUTPUT_DIR / "audios"
BOOK_LIST_FILE = DATA_DIR / "book_list_500.xlsx"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
logger = logging.getLogger(__name__)


# ==================== talelin API ====================
class TalelinBookAPI:
    """talelin图书搜索API - 免费无需key"""

    BASE_URL = "http://t.talelin.com/v2/book/search"

    @classmethod
    def search(cls, book_name: str, count: int = 1) -> Optional[Dict]:
        """搜索书籍，返回第一个结果"""
        try:
            url = f"{cls.BASE_URL}?q={url_quote(book_name)}&start=0&count={count}"
            with httpx.Client(timeout=30.0) as client:
                response = client.get(url)
                if response.status_code == 200:
                    data = response.json()
                    books = data.get("books", [])
                    if books:
                        return cls._normalize(books[0])
        except Exception as e:
            logger.warning(f"talelin搜索失败: {e}")
        return None

    @classmethod
    def _normalize(cls, book: Dict) -> Dict:
        """标准化API返回的数据"""
        author = book.get("author", [])
        if isinstance(author, list):
            author = ", ".join(author)

        images = book.get("images", {})
        cover_url = images.get("large") or images.get("medium") or book.get("image", "")

        return {
            "title": book.get("title", ""),
            "author": author,
            "publisher": book.get("publisher", ""),
            "isbn": book.get("isbn10") or book.get("isbn13", ""),
            "publishDate": book.get("pubdate", ""),
            "pages": book.get("pages", ""),
            "category": book.get("category", ""),
            "coverUrl": cover_url,
            "description": book.get("summary", ""),
        }


# ==================== MiniMax API ====================
class MiniMaxClient:
    """MiniMax API 客户端 - Anthropic兼容格式"""

    def __init__(self, api_key: str, base_url: str = "https://api.minimaxi.com/anthropic"):
        self.api_key = api_key
        self.base_url = base_url.rstrip('/')
        self.model = "MiniMax-M2.7"

    def chat(self, system_prompt: str, user_prompt: str, max_tokens: int = 8000,
             temperature: float = 0.7) -> str:
        """调用API生成内容"""
        url = f"{self.base_url}/v1/messages"
        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json",
            "anthropic-version": "2023-06-01"
        }
        payload = {
            "model": self.model,
            "max_tokens": max_tokens,
            "temperature": temperature,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ]
        }

        import httpx
        with httpx.Client(timeout=180.0) as client:
            response = client.post(url, headers=headers, json=payload)
            response.raise_for_status()
            result = response.json()
            # 返回 text 块的内容，如果没有则取 thinking 块
            for block in result.get("content", []):
                if block.get("type") == "text":
                    return block.get("text", "")
            # 如果没有 text 块，取最后一个块的文本（可能是 thinking）
            if result.get("content"):
                last_block = result["content"][-1]
                return last_block.get("text") or last_block.get("thinking", "")
            raise ValueError(f"API返回中未找到文本内容: {result}")


# ==================== 分类系统 ====================
CATEGORY_MAP = {
    1: "经典名著", 2: "儿童文学", 3: "科普百科",
    4: "历史传记", 5: "哲学心理", 6: "文学小说",
    7: "诗词歌赋", 8: "家庭教育", 9: "成长励志",
    10: "科幻悬疑", 11: "散文随笔", 12: "其他"
}

CATEGORY_KEYWORDS = {
    1: ["名著", "经典", "四大名著", "红楼梦", "三国", "水浒", "西游"],
    2: ["儿童", "童话", "绘本", "少年", "小王子", "小豆豆"],
    3: ["科普", "百科", "科学", "宇宙", "物理", "化学", "十万个"],
    4: ["历史", "传记", "史记", "人物", "帝王", "朝代"],
    5: ["哲学", "心理", "思维", "逻辑", "禅", "冥想", "瓦尔登湖"],
    6: ["小说", "文学", "故事", "虚构", "余华", "路遥", "莫言"],
    7: ["诗词", "诗歌", "唐诗", "宋词", "诗经", "古文"],
    8: ["教育", "家庭", "育儿", "管教", "妈妈", "亲子"],
    9: ["励志", "成长", "奋斗", "成功", "钢铁"],
    10: ["科幻", "悬疑", "推理", "三体", "东野", "侦探"],
    11: ["散文", "随笔", "杂文", "游记", "三毛", "鲁迅"],
    12: ["其他"]
}

CATEGORY_COLORS = {
    1: "#8B4513", 2: "#FF6B6B", 3: "#4ECDC4",
    4: "#9B59B6", 5: "#3498DB", 6: "#E74C3C",
    7: "#F39C12", 8: "#27AE60", 9: "#1ABC9C",
    10: "#2C3E50", 11: "#E67E22", 12: "#95A5A6"
}


def auto_classify(title: str, author: str = "", description: str = "") -> tuple:
    """根据书名/作者/简介自动推断分类"""
    text = f"{title} {author} {description}"
    for cat_id, keywords in CATEGORY_KEYWORDS.items():
        for kw in keywords:
            if kw in text:
                return cat_id, CATEGORY_MAP[cat_id]
    return 12, "其他"


# ==================== MiniMax 内容生成 ====================
def generate_ai_content(client: MiniMaxClient, title: str, author: str = "") -> Dict:
    """通过MiniMax生成：简介、金句、解读脚本"""

    # 1. 生成书籍简介
    intro_system = """你是一位资深的图书编辑。请严格按JSON格式输出，不要任何markdown标记。"""
    intro_prompt = f"""为《{title}》生成30字左右的简介，用于听书小程序展示。

输出格式（必须是有效JSON，不要任何其他文字）：
{{"intro": "简介内容"}}
"""
    try:
        intro_text = client.chat(intro_system, intro_prompt, max_tokens=500, temperature=0.7)
        intro_data = json.loads(intro_text)
        description = intro_data.get("intro", "")
    except Exception as e:
        logger.warning(f"简介生成失败: {e}")
        description = f"《{title}》是一本值得品读的经典之作"

    # 2. 生成10条金句
    quotes_system = """你是一位资深文学评论家。请严格按JSON格式输出，不要任何markdown标记。"""
    quotes_prompt = f"""为《{title}》提取10条最经典的金句，每条不超过50字。

输出格式（必须是有效JSON数组，不要任何其他文字）：
{{"quotes": ["金句1", "金句2", ...]}}
"""
    try:
        quotes_text = client.chat(quotes_system, quotes_prompt, max_tokens=2000, temperature=0.7)
        quotes_data = json.loads(quotes_text)
        quotes = quotes_data.get("quotes", [])
    except Exception as e:
        logger.warning(f"金句生成失败: {e}")
        quotes = []

    # 3. 生成约3000字的解读脚本（保留[停顿]标记用于edge-tts停顿效果）
    script_system = """你是一位资深文学解读专家。生成适合TTS朗读的脚本，必须包含[停顿X秒]等韵律标记。"""
    script_prompt = f"""为《{title}》{"作者" + author if author else "作者佚名"}生成一篇约3000字的深度解读脚本，适合TTS朗读。

要求：
1. 语言口语化，适合朗读
2. 每段结束时标注[停顿1秒]
3. 重要观点前标注[强调]
4. 纯文本输出，不要markdown

输出格式（必须是有效JSON，不要任何其他文字）：
{{"script": "脚本内容"}}
"""
    try:
        script_text = client.chat(script_system, script_prompt, max_tokens=8000, temperature=0.7)
        script_data = json.loads(script_text)
        script = script_data.get("script", "")
    except Exception as e:
        logger.warning(f"脚本生成失败: {e}")
        script = f"欢迎收听《{title}》解读。"

    return {
        "description": description,
        "quotes": quotes,
        "script": script
    }


# ==================== 字幕清理 ====================
def clean_subtitle_text(text: str) -> str:
    """清理字幕中的TTS标记，只用于显示"""
    # 移除 [停顿X秒] [停顿X分X秒]
    text = re.sub(r'\[停顿\d+(?:\.\d+)?(?:分)?(?:秒)?\]', '', text)
    # 移除 [强调]
    text = re.sub(r'\[强调\]', '', text)
    # 移除 [叹息] [欢呼] 等情感标记
    text = re.sub(r'\[(?:叹息|欢呼|兴奋|平静|低沉|高亢|温柔|激动)\]', '', text)
    # 清理多余空白
    text = re.sub(r'\n{3,}', '\n\n', text)
    text = text.strip()
    return text


def generate_srt_subtitle(script: str, output_path: Path):
    """根据脚本生成SRT字幕文件（清理版）"""
    # 这里需要实际生成字幕时间轴
    # 简化版：按字符数估算时间
    import math

    lines = []
    start = 0.0
    char_count = 0
    max_chars_per_line = 50
    max_duration = 5.0  # 每行最长5秒

    # 清理后的文本
    clean_text = clean_subtitle_text(script)

    # 按标点分割句子
    sentences = re.split(r'([。！？；\n])', clean_text)
    current_line = ""

    for i, part in enumerate(sentences):
        if i % 2 == 0:  # 文本部分
            current_line += part
        else:  # 标点部分
            current_line += part
            if len(current_line) >= max_chars_per_line or part in '。！？':
                # 计算时长（约0.3秒/字）
                duration = max(1.0, min(len(current_line) * 0.3, max_duration))
                start_ms = int(start * 1000)
                end_ms = int((start + duration) * 1000)

                lines.append(f"{len(lines) + 1}")
                lines.append(f"{ms_to_srt_time(start_ms)} --> {ms_to_srt_time(end_ms)}")
                lines.append(current_line.strip())
                lines.append("")

                start += duration + 0.5  # 加停顿
                current_line = ""

    # 处理剩余内容
    if current_line.strip():
        duration = max(1.0, len(current_line) * 0.3)
        start_ms = int(start * 1000)
        end_ms = int((start + duration) * 1000)
        lines.append(f"{len(lines) + 1}")
        lines.append(f"{ms_to_srt_time(start_ms)} --> {ms_to_srt_time(end_ms)}")
        lines.append(current_line.strip())
        lines.append("")

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    logger.info(f"字幕已生成: {output_path}")


def ms_to_srt_time(ms: int) -> str:
    """毫秒转SRT时间格式 HH:MM:SS,mmm"""
    s = ms // 1000
    ms = ms % 1000
    h = s // 3600
    m = (s % 3600) // 60
    s = s % 60
    return f"{h:02d}:{m:02d}:{s:02d},{ms:03d}"


# ==================== edge-tts 音频生成 ====================
async def generate_audio(text: str, output_path: Path, voice: str = "zh-CN-YunxiNeural") -> bool:
    """使用edge-tts生成音频（保留标记以产生真实停顿）"""
    try:
        import edge_tts
        AUDIO_DIR.mkdir(parents=True, exist_ok=True)

        # 限制文本长度（edge-tts 最大约10000字符）
        max_len = 8000
        if len(text) > max_len:
            logger.warning(f"文本过长({len(text)}字符)，截断到{max_len}字符")
            text = text[:max_len]

        communicate = edge_tts.Communicate(text, voice)
        await communicate.save(str(output_path))

        # 检查文件是否有效
        if output_path.exists() and output_path.stat().st_size > 0:
            logger.info(f"音频已生成: {output_path} ({output_path.stat().st_size} bytes)")
            return True
        else:
            logger.error("生成的文件为空")
            return False
    except Exception as e:
        logger.error(f"音频生成失败: {e}")
        return False


# ==================== 云端上传 ====================
def load_env():
    """加载环境变量"""
    for env_path in [Path(__file__).parent / ".env", BASE_DIR / ".env"]:
        if env_path.exists():
            with open(env_path, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#') and '=' in line:
                        key, value = line.split('=', 1)
                        os.environ.setdefault(key.strip(), value.strip())


def get_access_token() -> str:
    """获取微信access_token"""
    app_id = os.getenv("WECHAT_APP_ID")
    secret = os.getenv("WECHAT_SECRET")

    url = f"https://api.weixin.qq.com/cgi-bin/token?grant_type=client_credential&appid={app_id}&secret={secret}"
    with httpx.Client(timeout=30.0) as client:
        response = client.get(url)
        result = response.json()
        if "access_token" in result:
            return result["access_token"]
        raise RuntimeError(f"获取access_token失败: {result}")


def upload_to_cloud_storage(file_path: Path) -> Optional[str]:
    """上传文件到云存储，返回fileID（两步：1.获取链接 2.上传文件）"""
    try:
        access_token = get_access_token()
        env_id = os.getenv("WECHAT_ENV_ID")

        # 步骤1：获取上传链接
        url = f"https://api.weixin.qq.com/tcb/uploadfile?access_token={access_token}"
        data = {
            "env": env_id,
            "path": f"audio/{file_path.name}"
        }

        with httpx.Client(timeout=30.0) as client:
            response = client.post(url, json=data)
            result = response.json()
            logger.info(f"获取上传链接响应: {result}")

            if result.get("errcode") != 0:
                logger.error(f"获取上传链接失败: {result}")
                return None

            upload_url = result.get("url")
            token = result.get("token")
            authorization = result.get("authorization")
            cos_file_id = result.get("cos_file_id")
            file_id = result.get("file_id")

        # 步骤2：上传实际文件
        files = {
            'file': (file_path.name, open(file_path, 'rb'), 'audio/mpeg')
        }
        data = {
            'key': f"audio/{file_path.name}",
            'Signature': authorization,
            'x-cos-security-token': token,
            'x-cos-meta-fileid': cos_file_id,
        }

        with httpx.Client(timeout=120.0) as client:
            response = client.post(upload_url, files=files, data=data)
            if response.status_code in [200, 201, 204]:
                logger.info(f"✅ 文件上传成功: {file_id}")
                return file_id
            else:
                logger.error(f"文件上传失败: {response.status_code} - {response.text}")
                return None

    except Exception as e:
        logger.error(f"云存储上传失败: {e}")
        return None


def check_cloud_exists(title: str) -> bool:
    """检查云数据库是否已存在该书籍"""
    try:
        access_token = get_access_token()
        env_id = os.getenv("WECHAT_ENV_ID")

        url = f"https://api.weixin.qq.com/tcb/databasequery?access_token={access_token}"
        query = f"""
db.collection('books').where({{title: '{title}'}}).limit(1).get()
"""
        data = {"env": env_id, "query": query.strip()}

        with httpx.Client(timeout=30.0) as client:
            response = client.post(url, json=data)
            result = response.json()

            if result.get("errcode") == 0:
                data_list = result.get("data", [])
                return len(data_list) > 0
        return False
    except Exception as e:
        logger.warning(f"云端查询失败: {e}")
        return False


def import_to_cloud_db(book_data: Dict) -> bool:
    """导入书籍数据到云数据库"""
    try:
        access_token = get_access_token()
        env_id = os.getenv("WECHAT_ENV_ID")

        # 构建字段表达式 - 正确处理各种数据类型
        fields = []
        for key, value in book_data.items():
            if value is None or value == "":
                # 空值用null
                fields.append(f"{key}: null")
            elif isinstance(value, bool):
                fields.append(f"{key}: {str(value).lower()}")
            elif isinstance(value, (int, float)):
                fields.append(f"{key}: {value}")
            elif isinstance(value, str):
                # 正确转义: 先转义反斜杠，再转义双引号
                escaped = value.replace('\\', '\\\\').replace('"', '\\"')
                fields.append(f'{key}: "{escaped}"')
            elif isinstance(value, list):
                # 处理数组 - 每个元素都是字符串金句
                items = []
                for item in value:
                    if isinstance(item, str):
                        item_escaped = item.replace('\\', '\\\\').replace('"', '\\"')
                        items.append(f'"{item_escaped}"')
                    else:
                        items.append(str(item))
                fields.append(f"{key}: [{', '.join(items)}]")
            else:
                # 其他类型转为字符串
                fields.append(f'{key}: "{str(value)}"')

        data_str = ", ".join(fields)
        query = 'db.collection("books").add({data: {' + data_str + '}})'
        
        logger.info(f"导入数据查询: {query[:200]}...")

        url = f"https://api.weixin.qq.com/tcb/databaseadd?access_token={access_token}"
        data = {
            "env": env_id,
            "query": query
        }

        logger.info(f"导入数据查询: {query[:200]}...")

        with httpx.Client(timeout=30.0) as client:
            response = client.post(url, json=data)
            result = response.json()

            if result.get("errcode") == 0:
                logger.info(f"✅ 已导入云数据库: {book_data.get('title')}")
                return True
            else:
                logger.error(f"导入失败: {result}")
        return False
    except Exception as e:
        logger.error(f"云数据库导入失败: {e}")
        return False


# ==================== Excel书单管理 ====================
def load_book_list() -> List[Dict]:
    """加载Excel书单"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(BOOK_LIST_FILE)
        ws = wb.active

        books = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # 书名
                books.append({
                    "书名": row[0],
                    "是否导入": row[1] or "否",
                    "封面URL": row[2] or "",
                    "分类": row[3] or ""
                })
        return books
    except Exception as e:
        logger.error(f"加载书单失败: {e}")
        return []


def update_book_status(book_name: str, cover_url: str = ""):
    """更新书单中书籍的状态"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(BOOK_LIST_FILE)
        ws = wb.active

        for row in ws.iter_rows(min_row=2):
            if row[0].value == book_name:
                row[1].value = "是"  # 是否导入
                if cover_url:
                    row[2].value = cover_url  # 封面URL
                break

        wb.save(BOOK_LIST_FILE)
        logger.info(f"✅ 已更新书单状态: {book_name}")
    except Exception as e:
        logger.error(f"更新书单失败: {e}")


# ==================== 主流程 ====================
def process_book(book_name: str, minimax_client: MiniMaxClient, force: bool = False) -> bool:
    """处理单本书籍的完整流程"""
    logger.info(f"\n{'='*50}")
    logger.info(f"📖 开始处理: 《{book_name}》")
    logger.info(f"{'='*50}")

    # 1. 检查是否已处理
    if not force and check_cloud_exists(book_name):
        logger.info(f"⏭️ 云端已存在: 《{book_name}》，跳过")
        return False

    # 2. 获取书籍信息（talelin API）
    logger.info(f"[1/5] 通过talelin API获取书籍信息...")
    talelin_info = TalelinBookAPI.search(book_name)
    if talelin_info:
        logger.info(f"   ✅ 获取成功: 封面URL = {talelin_info.get('coverUrl', 'N/A')[:50]}...")
        cover_url = talelin_info.get("coverUrl", "")
        author = talelin_info.get("author", "")
        publisher = talelin_info.get("publisher", "")
        isbn = talelin_info.get("isbn", "")
    else:
        logger.warning(f"   ⚠️ talelin未找到，使用默认值")
        cover_url = ""
        author = ""
        publisher = ""
        isbn = ""

    # 3. AI生成内容（MiniMax）
    logger.info(f"[2/5] 通过MiniMax生成内容...")
    ai_content = generate_ai_content(minimax_client, book_name, author)
    logger.info(f"   ✅ 简介: {ai_content['description'][:30]}...")
    logger.info(f"   ✅ 金句: {len(ai_content['quotes'])}条")
    logger.info(f"   ✅ 脚本: {len(ai_content['script'])}字符")

    # 4. 自动分类
    cat_id, cat_name = auto_classify(book_name, author, ai_content.get("description", ""))
    cover_color = CATEGORY_COLORS.get(cat_id, "#95A5A6")

    # 5. 生成音频（edge-tts）
    logger.info(f"[3/5] 生成音频...")
    audio_path = AUDIO_DIR / f"{book_name}.mp3"
    asyncio.run(generate_audio(ai_content["script"], audio_path))

    # 6. 生成字幕
    logger.info(f"[4/5] 生成字幕...")
    subtitle_path = AUDIO_DIR / f"{book_name}.srt"
    generate_srt_subtitle(ai_content["script"], subtitle_path)

    # 7. 上传云存储
    logger.info(f"[4.5/5] 上传到云存储...")
    file_id = upload_to_cloud_storage(audio_path) if audio_path.exists() else ""

    # 8. 准备云数据库数据
    book_data = {
        "title": book_name,
        "author": author,
        "category": cat_id,
        "categoryName": cat_name,
        "publisher": publisher,
        "isbn": isbn,
        "publishDate": "",
        "pages": 0,
        "description": ai_content["description"],
        "coverUrl": cover_url,
        "coverColor": cover_color,
        "script": ai_content["script"],  # 原始脚本含停顿标记
        "cleanScript": clean_subtitle_text(ai_content["script"]),  # 清理版
        "quotes": ai_content["quotes"],
        "audioFileId": file_id,
        "audioUrl": f"cloud://{os.getenv('WECHAT_ENV_ID')}/audio/{book_name}.mp3",
        "updateTime": int(time.time() * 1000)
    }

    # 9. 导入云数据库
    logger.info(f"[5/5] 导入云数据库...")
    success = import_to_cloud_db(book_data)

    if success:
        # 10. 更新书单状态
        update_book_status(book_name, cover_url)
        logger.info(f"✅ 《{book_name}》处理完成!")
    else:
        logger.error(f"❌ 《{book_name}》导入失败")

    return success


def main():
    parser = argparse.ArgumentParser(description="听书小程序全自动Pipeline")
    parser.add_argument("--count", type=int, default=1, help="处理书籍数量")
    parser.add_argument("--book", type=str, default="", help="指定书籍名称")
    parser.add_argument("--force", action="store_true", help="强制重新处理")
    args = parser.parse_args()

    # 加载配置
    load_env()
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        logger.error("❌ 未配置 ANTHROPIC_API_KEY")
        return

    minimax_client = MiniMaxClient(api_key, os.getenv("ANTHROPIC_BASE_URL", "https://api.minimaxi.com/anthropic"))

    # 确保输出目录存在
    AUDIO_DIR.mkdir(parents=True, exist_ok=True)

    if args.book:
        # 指定书籍
        process_book(args.book, minimax_client, args.force)
    else:
        # 从书单选取
        books = load_book_list()
        pending = [b for b in books if b["是否导入"] != "是"]

        if not pending:
            logger.info("📚 所有书籍已处理完成!")
            return

        count = min(args.count, len(pending))
        logger.info(f"📚 待处理: {len(pending)}本 | 本次处理: {count}本")

        for i, book in enumerate(pending[:count], 1):
            logger.info(f"\n>>> 进度: {i}/{count}")
            try:
                process_book(book["书名"], minimax_client, args.force)
            except Exception as e:
                logger.error(f"处理失败: {e}")
                continue


if __name__ == "__main__":
    main()
