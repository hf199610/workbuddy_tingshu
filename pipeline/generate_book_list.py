#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
生成热门书籍列表
输出：Excel格式（书名 | 是否导入 | 封面URL）
"""

import os
import json
import logging
from pathlib import Path
from datetime import datetime

# 路径配置
BASE_DIR = Path(__file__).parent.parent
DATA_DIR = BASE_DIR / "data_source"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
logger = logging.getLogger(__name__)


# ==================== 分类书单模板 ====================
# 按分类预设一部分知名书籍，便于生成完整列表

CATEGORY_BOOKS = {
    "经典名著": [
        "红楼梦", "三国演义", "水浒传", "西游记", "活着", "围城", "平凡的世界",
        "白鹿原", "四世同堂", "骆驼祥子", "红高粱", "蛙", "丰乳肥臀", "生死场",
        "呼啸山庄", "傲慢与偏见", "战争与和平", "百年孤独", "追风筝的人", "了不起的盖茨比",
        "巴黎圣母院", "悲惨世界", "老人与海", "飘", "简爱", "茶花女", "苔丝",
        "源氏物语", "挪威的森林", "解忧杂货店", "刺杀骑士团长",
    ],
    "儿童文学": [
        "小王子", "窗边的小豆豆", "夏洛的网", "绿野仙踪", "爱丽丝梦游仙境",
        "长袜子皮皮", "小猪唏哩呼噜", "没头脑和不高兴", "笨狼的故事", "狼王梦",
        "草房子", "青铜葵花", "根鸟", "山羊不吃天堂草", "蜻蜓眼",
        "猜猜我有多爱你", "逃家小兔", "爷爷一定有办法", "棕色的熊你在看什么",
        "好饿的毛毛虫", "大卫不可以", "我的爸爸叫焦尼", "爷爷变成了幽灵",
    ],
    "科普百科": [
        "时间简史", "人类简史", "未来简史", "万物简史", "从一到无穷大",
        "自私的基因", "盲眼钟表匠", "生命是什么", "宇宙的琴弦", "时间之箭",
        "皇帝新脑", "果壳中的宇宙", "大设计", "宇宙的结构", "现实不似你所见",
        "什么是数学", "物理学的进化", "伽利略的手指", "上帝掷骰子吗",
        "寂静的春天", "只有一个地球", "昆虫记", "物种起源", "自私的基因",
    ],
    "历史传记": [
        "史记", "资治通鉴", "万历十五年", "明朝那些事儿", "半小时漫画中国史",
        "全球通史", "人类的故事", "耶路撒冷三千年", "奥斯维辛一部历史",
        "习近平谈治国理政", "毛泽东选集", "邓小平时代", "曾国藩传", "张作霖传",
        "苏东坡传", "王阳明传", "杜甫传", "李时珍传", "富兰克林自传",
        "史蒂夫乔布斯传", "埃隆马斯克传", "曾国藩家书", "FBI教你读心术",
    ],
    "哲学心理": [
        "瓦尔登湖", "沉思录", "人生的智慧", "幸福之路", "自卑与超越",
        "梦的解析", "精神分析引论", "乌合之众", "自卑与超越", "当下的力量",
        "少有人走的路", "活出生命的意义", "幸福的婚姻", "为何家会伤人",
        "论语", "道德经", "庄子", "大学", "中庸", "孟子", "荀子",
        "心经", "金刚经", "六祖坛经", "苏菲的世界", "西方哲学史",
    ],
    "文学小说": [
        "三体", "流浪地球", "基地", "银河帝国", "海伯利安", "太空漫游",
        "解忧杂货店", "白夜行", "嫌疑人X的献身", "恶意", "新参者", "放学后",
        "挪威的森林", "1Q84", "海边的卡夫卡", "世界尽头与冷酷仙境",
        "繁花", "推拿", "蛙", "丰乳肥臀", "生死疲劳", "天堂向左地狱向右",
        "盗墓笔记", "鬼吹灯", "明朝那些事儿", "琅琊榜", "三生三世十里桃花",
        "小时代", "爵迹", "长安十二时辰", "长安的荔枝", "太白金星有点烦",
    ],
    "诗词歌赋": [
        "唐诗三百首", "宋词三百首", "诗经", "楚辞", "元曲三百首",
        "古文观止", "唐宋八大家散文", "千古词帝李煜", "纳兰词", "李清照词传",
        "李白诗传", "杜甫诗传", "白居易诗传", "苏轼词传", "辛弃疾词传",
        "仓央嘉措诗集", "汪国真诗选", "余光中诗选", "舒婷诗选", "北岛诗选",
        "海子的诗", "面朝大海春暖花开", "飞鸟集", "新月集", "泰戈尔诗选",
    ],
    "家庭教育": [
        "好妈妈胜过好老师", "正面管教", "如何说孩子才会听怎么听孩子才肯说",
        "非暴力沟通", "养育男孩", "养育女孩", "童年的秘密", "捕捉孩子的敏感期",
        "蒙特梭利家庭方案", "德国幼儿教育", "犹太人教子枕边书",
        "傅雷家书", "梁启超家书", "曾国藩家书", "颜氏家训", "朱子家训",
        "最美的教育最简单", "让孩子的大脑自由", "全脑教养法", "正面管教",
    ],
    "成长励志": [
        "钢铁是怎样炼成的", "鲁滨逊漂流记", "老人与海", "假如给我三天光明",
        "羊皮卷", "人性的弱点", "思考致富", "唤醒内心的巨人", "世界上最伟大的推销员",
        "肖申克的救赎", "阿甘正传", "当幸福来敲门", "风雨哈佛路", "硅谷钢铁侠",
        "认知觉醒", "刻意练习", "高效能人士的七个习惯", "深度工作",
        "原则", "债务危机", "大空头", "穷爸爸富爸爸", "小狗钱钱",
    ],
    "科幻悬疑": [
        "三体", "三体Ⅱ", "三体Ⅲ", "流浪地球", "超新星纪元", "球状闪电",
        "乡村教师", "朝闻道", "带上她的眼睛", "地球大炮", "镜子", "命运",
        "基地", "基地与帝国", "第二基地", "基地前传", "基地后传",
        "海伯利安", "海伯利安的陨落", "安迪密恩", "安迪密恩的觉醒",
        "银河帝国", "太空漫游", "永恒的终结", "神门自己", "日暮",
        "无人生还", "东方快车谋杀案", "尼罗河上的惨案", "ABC谋杀案",
        "福尔摩斯探案集", "白夜行", "嫌疑人X的献身", "恶意", "红手指",
    ],
    "散文随笔": [
        "目送", "孩子你慢慢来", "亲爱的安德烈", "我们仨", "干校六记",
        "背影", "荷塘月色", "故都的秋", "湘行散记", "边城",
        "瓦尔登湖", "我在雨中等你", "当我谈跑步时我谈些什么", "小王子",
        "三毛流浪记", "撒哈拉的故事", "哭泣的骆驼", "万水千山走遍",
        "林清玄散文", "张晓风散文", "毕淑敏散文", "周国平散文", "汪曾祺散文",
        "鲁迅杂文", "梁实秋散文", "丰子恺散文", "杨绛散文", "季羡林散文",
    ],
    "其他": [
        "道德经", "易经", "黄帝内经", "山海经", "世说新语",
        "搜神记", "聊斋志异", "阅微草堂笔记", "子不语", "列子",
        "梦溪笔谈", "天工开物", "农政全书", "本草纲目", "徐霞客游记",
    ]
}


def expand_book_list(target_count: int = 500) -> list:
    """扩展书籍列表到指定数量"""
    all_books = []
    book_set = set()

    # 按分类顺序添加书籍，保持多样性
    categories = list(CATEGORY_BOOKS.keys())

    # 第一轮：每个分类取一部分，均匀分布
    base_per_category = target_count // len(categories)

    for category, books in CATEGORY_BOOKS.items():
        # 每本书后加特殊标记避免重复
        for i, book in enumerate(books):
            key = f"{book}_{i}"  # 处理重名
            if key not in book_set:
                book_set.add(key)
                all_books.append({
                    "书名": book,
                    "是否导入": "否",
                    "封面URL": "",
                    "分类": category
                })

    # 如果不够，生成变体
    if len(all_books) < target_count:
        need_more = target_count - len(all_books)
        variants = [
            "全集", "精选", "插图本", "注音版", "英文版", "插画版",
            "青少版", "典藏版", "豪华版", "精装版", "普及版", "图解版",
        ]
        extra_books = []
        for cat_name, books in CATEGORY_BOOKS.items():
            for book in books:
                for variant in variants:
                    if len(extra_books) >= need_more:
                        break
                    new_book = f"{book}{variant}"
                    key = f"{new_book}_{cat_name}"
                    if key not in book_set:
                        book_set.add(key)
                        extra_books.append({
                            "书名": new_book,
                            "是否导入": "否",
                            "封面URL": "",
                            "分类": cat_name
                        })
            if len(extra_books) >= need_more:
                break

        all_books.extend(extra_books[:need_more])

    return all_books[:target_count]


def save_to_excel(books: list, output_path: Path):
    """保存书籍列表到Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "书籍列表"

        # 表头样式
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写入表头
        headers = ["书名", "是否导入", "封面URL", "分类"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入数据
        for row, book in enumerate(books, 2):
            ws.cell(row=row, column=1, value=book["书名"]).border = thin_border
            ws.cell(row=row, column=2, value=book["是否导入"]).border = thin_border
            ws.cell(row=row, column=3, value=book["封面URL"]).border = thin_border
            ws.cell(row=row, column=4, value=book["分类"]).border = thin_border

        # 设置列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15

        wb.save(output_path)
        logger.info(f"✅ 已保存 {len(books)} 本书籍到 {output_path}")
        return True

    except ImportError:
        logger.error("❌ 需要安装 openpyxl: pip install openpyxl")
        # 降级为CSV
        import csv
        with open(output_path.with_suffix('.csv'), 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=["书名", "是否导入", "封面URL", "分类"])
            writer.writeheader()
            writer.writerows(books)
        logger.info(f"✅ 已保存 {len(books)} 本书籍到 {output_path.with_suffix('.csv')}")
        return True


def main():
    output_file = DATA_DIR / "book_list_500.xlsx"

    logger.info("📚 开始生成热门500本书列表...")

    books = expand_book_list(500)
    save_to_excel(books, output_file)

    # 统计
    category_count = {}
    for book in books:
        cat = book["分类"]
        category_count[cat] = category_count.get(cat, 0) + 1

    logger.info("📊 分类统计:")
    for cat, count in category_count.items():
        logger.info(f"  {cat}: {count}本")


if __name__ == "__main__":
    main()
