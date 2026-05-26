#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""生成Excel版本的书籍数据备份"""

import csv
import os
from pathlib import Path
from datetime import datetime

# Try to use openpyxl, fallback to xlsxwriter
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYL = True
except ImportError:
    HAS_OPENPYL = False

PIPELINE_DIR = Path(__file__).parent
CSV_FILE = 'D:/小程序开发/workbud_tingshu/data_source/book_list_500_enhanced.csv'
OUTPUT_DIR = PIPELINE_DIR / "output"

def fix_duplicate_sishitang():
    """修复四世同堂重复ID问题"""
    import httpx
    import os
    
    # Load env
    for env_path in [PIPELINE_DIR / ".env", PIPELINE_DIR.parent / ".env"]:
        if env_path.exists():
            with open(env_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#") and "=" in line:
                        key, value = line.split("=", 1)
                        os.environ.setdefault(key.strip(), value.strip())
    
    app_id = os.environ.get("WECHAT_APP_ID")
    secret = os.environ.get("WECHAT_SECRET")
    
    # Get access token
    resp = httpx.get(f"https://api.weixin.qq.com/cgi-bin/token?grant_type=client_credential&appid={app_id}&secret={secret}", timeout=30)
    access_token = resp.json()["access_token"]
    env_id = os.environ.get("WECHAT_ENV_ID")
    
    # Query for 四世同堂
    query = 'db.collection("books").where({title: "四世同堂"}).get()'
    resp = httpx.post(f"https://api.weixin.qq.com/tcb/databasequery?access_token={access_token}",
                      json={"env": env_id, "query": query}, timeout=30)
    result = resp.json()
    
    if result.get("errcode") == 0 and result.get("data"):
        ids = []
        for item in result.get("data", []):
            record = item
            if isinstance(item, str):
                import json
                try:
                    record = json.loads(item)
                except:
                    pass
            record_id = record.get('_id') or record.get('id')
            ids.append(record_id)
        
        print(f"Found {len(ids)} 四世同堂 records: {ids}")
        
        # Delete duplicates (keep first)
        if len(ids) > 1:
            for record_id in ids[1:]:
                del_query = f'db.collection("books").doc("{record_id}").remove()'
                resp = httpx.post(f"https://api.weixin.qq.com/tcb/databasedelete?access_token={access_token}",
                                  json={"env": env_id, "query": del_query}, timeout=30)
                print(f"Deleted duplicate: {record_id}, result: {resp.json()}")
            
            # Update CSV with correct ID
            correct_id = ids[0]
            rows = []
            with open(CSV_FILE, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                header = next(reader)
                rows.append(header)
                for row in reader:
                    if len(row) > 0 and row[0] == '四世同堂':
                        row[8] = correct_id
                    rows.append(row)
            with open(CSV_FILE, 'w', encoding='utf-8', newline='') as f:
                writer = csv.writer(f)
                writer.writerows(rows)
            print(f"Updated CSV with correct ID: {correct_id}")

def generate_excel():
    """生成Excel备份"""
    if not HAS_OPENPYL:
        print("openpyxl not available, skipping Excel generation")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "书籍列表"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["书名", "作者", "是否导入", "分类", "音频FileID", "讲解稿", "金句", "导入时间", "数据库ID", "错误信息", "金句数", "时长"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data
    with open(CSV_FILE, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader)  # skip header
        for row_idx, row in enumerate(reader, 2):
            while len(row) < 12:
                row.append('')
            for col, value in enumerate(row[:12], 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
    
    # Auto adjust column width
    for col in range(1, 13):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    # Save
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = OUTPUT_DIR / f"book_list_backup_{timestamp}.xlsx"
    wb.save(excel_file)
    print(f"Excel saved: {excel_file}")

def main():
    print("=" * 60)
    print("生成Excel备份并修复重复记录")
    print("=" * 60)
    
    # Fix duplicate
    fix_duplicate_sishitang()
    
    # Generate Excel
    generate_excel()
    
    print("=" * 60)
    print("完成!")
    print("=" * 60)

if __name__ == "__main__":
    main()